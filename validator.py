import os
import json
import pandas as pd
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter


# =====================================================
# CLEAN FUNCTIONS
# =====================================================

def extract_entry_time(dt_str):
    try:
        entry_line = dt_str.split('\n')[1]
        return datetime.strptime(entry_line.strip(), "%b %d, %Y, %H:%M")
    except:
        return None


def extract_entry_signal(signal_str):
    try:
        return signal_str.split('\n')[1].strip()
    except:
        return None


def extract_trade_type(type_str):
    try:
        return type_str.split('\n')[1].strip()
    except:
        return None


# =====================================================
# EXCEL SAVE FUNCTION
# =====================================================

def save_excel(df, path):

    df.to_excel(path, index=False)

    wb = load_workbook(path)
    ws = wb.active

    ws.freeze_panes = "A2"

    for column in ws.columns:
        max_length = 0
        col_letter = get_column_letter(column[0].column)

        for cell in column:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))

        ws.column_dimensions[col_letter].width = max_length + 3

    for cell in ws[1]:
        cell.font = Font(bold=True)

    wb.save(path)


# =====================================================
# MAIN VALIDATION FUNCTION
# =====================================================

def run_validation(ce_data, pe_data, index_data):

    base_dir = "validation_output"
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    base_dir = os.path.join(base_dir, f"validation_{timestamp}")

    valid_dir = os.path.join(base_dir, "valid")
    not_valid_dir = os.path.join(base_dir, "not_valid")

    os.makedirs(valid_dir, exist_ok=True)
    os.makedirs(not_valid_dir, exist_ok=True)

    df_ce = pd.DataFrame(ce_data)
    df_pe = pd.DataFrame(pe_data)
    df_index = pd.DataFrame(index_data)

    for df in [df_ce, df_pe, df_index]:
        df['entry_time'] = df['dateTime'].apply(extract_entry_time)
        df['entry_signal'] = df['signal'].apply(extract_entry_signal)
        df['trade_type'] = df['type'].apply(extract_trade_type)
        df.dropna(subset=['entry_time'], inplace=True)

    df_ce = df_ce[df_ce['trade_type'].str.contains('Entry', na=False)]
    df_pe = df_pe[df_pe['trade_type'].str.contains('Entry', na=False)]
    df_index = df_index[df_index['trade_type'].str.contains('Entry', na=False)]

    df_ce = df_ce.sort_values('entry_time')
    df_pe = df_pe.sort_values('entry_time')
    df_index = df_index.sort_values('entry_time')

    matched = []
    ce_unmatched = []
    pe_used = set()

    for _, ce in df_ce.iterrows():

        ce_time = ce['entry_time']
        ce_signal = ce['entry_signal']

        if ce_signal == "BUY":
            required_pe_signal = "SELL"
            required_index_signal = "BUY"
            confirmation_type = "CE Bullish Confirmed"

        elif ce_signal == "SELL":
            required_pe_signal = "BUY"
            required_index_signal = "SELL"
            confirmation_type = "CE Bearish Confirmed"

        else:
            continue

        pe_candidates = df_pe[
            (df_pe['entry_signal'] == required_pe_signal) &
            (abs(df_pe['entry_time'] - ce_time) <= pd.Timedelta('1min'))
        ]

        index_candidates = df_index[
            (df_index['entry_signal'] == required_index_signal) &
            (abs(df_index['entry_time'] - ce_time) <= pd.Timedelta('1min'))
        ]

        if not pe_candidates.empty and not index_candidates.empty:

            pe = pe_candidates.iloc[0]
            index_row = index_candidates.iloc[0]

            pe_used.add(pe.name)

            matched.append({

                "CE Symbol": ce.get('symbol'),
                "CE TradeNo": ce.get('tradeNo'),
                "CE Signal": ce_signal,
                "CE Time": ce_time.strftime("%d-%m-%Y %H:%M:%S"),

                "PE Symbol": pe.get('symbol'),
                "PE TradeNo": pe.get('tradeNo'),
                "PE Signal": required_pe_signal,
                "PE Time": pe['entry_time'].strftime("%d-%m-%Y %H:%M:%S"),

                "INDEX Symbol": index_row.get('symbol'),
                "INDEX TradeNo": index_row.get('tradeNo'),
                "INDEX Signal": required_index_signal,
                "INDEX Time": index_row['entry_time'].strftime("%d-%m-%Y %H:%M:%S"),

                "Confirmation Type": confirmation_type,
                "Status": "VALID"
            })

        else:
            ce_unmatched.append({
                "CE Symbol": ce.get('symbol'),
                "CE TradeNo": ce.get('tradeNo'),
                "Signal": ce_signal,
                "Time": ce_time.strftime("%d-%m-%Y %H:%M:%S"),
                "Reason": "PE or INDEX confirmation missing"
            })

    # =============================
    # PE UNMATCHED
    # =============================

    pe_unmatched = df_pe[~df_pe.index.isin(pe_used)].copy()
    pe_unmatched['Time'] = pe_unmatched['entry_time'].dt.strftime("%d-%m-%Y %H:%M:%S")
    pe_unmatched['Reason'] = "No CE confirmation"

    pe_unmatched['PE Symbol'] = pe_unmatched['symbol']

    pe_unmatched = pe_unmatched[['PE Symbol','tradeNo','entry_signal','trade_type','Time','Reason']]
    pe_unmatched.columns = ['PE Symbol','PE TradeNo','Signal','Trade Type','Time','Reason']

    # =============================
    # EXPORT FILES
    # =============================

    matched_df = pd.DataFrame(matched)

    save_excel(matched_df, os.path.join(valid_dir, "matched_signals.xlsx"))
    save_excel(pd.DataFrame(ce_unmatched), os.path.join(not_valid_dir, "ce_unmatched.xlsx"))
    save_excel(pe_unmatched, os.path.join(not_valid_dir, "pe_unmatched.xlsx"))

    # =============================
    # EXPORT MATCHED JSON
    # =============================

    matched_json_path = os.path.join(valid_dir, "matched_signals.json")

    if not matched_df.empty:
        matched_df.to_json(matched_json_path, orient="records", indent=4)
    else:
        with open(matched_json_path, "w") as f:
            f.write("[]")

    # =============================
    # META JSON FILE
    # =============================

    meta_json_path = os.path.join(base_dir, "validation_meta.json")

    meta_data = {
        "generated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "ce_symbols": sorted(df_ce['symbol'].dropna().unique().tolist()) if not df_ce.empty else [],
        "pe_symbols": sorted(df_pe['symbol'].dropna().unique().tolist()) if not df_pe.empty else [],
        "index_symbols": sorted(df_index['symbol'].dropna().unique().tolist()) if not df_index.empty else [],
        "ce_min_time": df_ce['entry_time'].min().strftime("%d-%m-%Y %H:%M:%S") if not df_ce.empty else None,
        "ce_max_time": df_ce['entry_time'].max().strftime("%d-%m-%Y %H:%M:%S") if not df_ce.empty else None,
        "pe_min_time": df_pe['entry_time'].min().strftime("%d-%m-%Y %H:%M:%S") if not df_pe.empty else None,
        "pe_max_time": df_pe['entry_time'].max().strftime("%d-%m-%Y %H:%M:%S") if not df_pe.empty else None,
        "index_min_time": df_index['entry_time'].min().strftime("%d-%m-%Y %H:%M:%S") if not df_index.empty else None,
        "index_max_time": df_index['entry_time'].max().strftime("%d-%m-%Y %H:%M:%S") if not df_index.empty else None,
        "total_ce_entries": len(df_ce),
        "total_pe_entries": len(df_pe),
        "total_index_entries": len(df_index),
        "total_valid_matches": len(matched_df)
    }

    with open(meta_json_path, "w") as f:
        json.dump(meta_data, f, indent=4)

    # =============================
    # GLOBAL SUMMARY
    # =============================

    summary = pd.DataFrame({
        "Metric": [
            "Total CE Entries",
            "Total PE Entries",
            "Total INDEX Entries",
            "Valid Trades",
            "CE Not Confirmed",
            "PE Not Confirmed",
            "Match Percentage"
        ],
        "Value": [
            len(df_ce),
            len(df_pe),
            len(df_index),
            len(matched_df),
            len(ce_unmatched),
            len(pe_unmatched),
            round((len(matched_df)/len(df_ce))*100,2) if len(df_ce)>0 else 0
        ]
    })

    save_excel(summary, os.path.join(base_dir, "summary.xlsx"))

    # =============================
    # RETURN PATHS
    # =============================

    return {
        "base_directory": base_dir,
        "matched_json_file": matched_json_path,
        "meta_json_file": meta_json_path
    }